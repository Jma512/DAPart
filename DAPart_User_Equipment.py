import base64
import json
import os
import pickle
import time
import openpyxl as op
import numpy as np
import requests
import torch
import sys
import subprocess
import re
import random
import multiprocessing

from subprocess import PIPE
from PIL import Image
from io import BytesIO
from torchvision import transforms
from policy_gradient_mj import REINFORCEAgent, ReplayBuffer
from vgg16 import VGG16
from mobilenetv2 import MobileNetV2
from resnet50 import ResNet50

# dict = {'VGG16': 17, 'AlexNet': 8, 'MobileNetV2': 21, 'ResNet50': 18, 'MobileViT': 13}

# 创建强化学习代理
state_size = 3
action_size = 18  # 替换为您自己的动作空间大小
agent = REINFORCEAgent(state_size, action_size)
server_ip = "192.168.0.103"  # 服务器ip
server_port = "8022"  # 服务器端口


class Util0():

    def __init__(self):
        self.id = 0
        self.timelist = []

    def settimelist(self, time1):
        dict = {}
        dict["id"] = self.id
        dict["time1"] = time1
        self.id = self.id + 1
        self.timelist.append(dict)

    def gettimelist(self):
        return self.timelist


class Util():

    def __init__(self):
        self.id = 0
        self.timelist = []

    def settimelist(self, time1, time2):
        dict = {}
        dict["id"] = self.id
        dict["time1"] = time1
        dict["time2"] = time2
        self.id = self.id + 1
        self.timelist.append(dict)

    def gettimelist(self):
        return self.timelist


class Util2():

    def __init__(self):
        self.id = 0
        self.timelist = []

    def settimelist(self, time1):
        dict = {}
        dict["id"] = self.id
        dict["time1"] = time1
        self.id = self.id + 1
        self.timelist.append(dict)

    def gettimelist(self):
        return self.timelist


# 创建子进程类，用于执行 tegrastats 命令
class TegraStatsProcess(multiprocessing.Process):
    def __init__(self):
        super().__init__()
        self.stdout_queue = multiprocessing.Queue()  # 输出队列
        self.stop_event = multiprocessing.Event()  # 停止信号

    def run(self):
        while not self.stop_event.is_set():
            # 执行 tegrastats 命令并将输出重定向到队列中
            command = ["tegrastats", "--interval", str(100)]
            with subprocess.Popen(command, stdout=subprocess.PIPE, universal_newlines=True) as proc:
                for line in proc.stdout:
                    self.stdout_queue.put(line.strip())
                    if self.stop_event.is_set():  # 如果接收到停止信号，则结束进程
                        break

    def stop(self):
        self.stop_event.set()


def get_upload_speed():
    url = f'http://{server_ip}:{server_port}/receivetestspeed'
    # 生成一个1Mb的虚拟文件
    file_content = b"H" * 1024 * 32
    file_size = len(file_content) * 8 / 1024  # 计算文件大小 单位kb
    virtual_file = BytesIO(file_content)

    try:
        start_time = time.time()
        # 发送POST请求上传文件
        response = requests.post(url, data=virtual_file, timeout=5)
        end_time = time.time()

        upload_time = end_time - start_time  # 计算上传所需的时间
        upload_speed = file_size / upload_time / 128  # 计算上传速度 单位Mbps
        print(upload_time)
        print(f"上传速度: {upload_speed:.2f} Mbps")
        return upload_speed
    except requests.exceptions.Timeout:
        ex_speed = 0
        print(f"EX速度: {ex_speed:.2f} Mbps")
        return ex_speed
    except requests.exceptions.RequestException as e:
        ex_speed = 0
        print(f"EX速度: {ex_speed:.2f} Mbps")
        return ex_speed


def get_signal_level_iwlist():
    try:
        command = 'iwlist wlan0 scanning | grep "Signal level"'
        result = subprocess.run(command, shell=True, stdout=PIPE, stderr=PIPE)
        output = result.stdout.strip()
        a = output[8:10]
        b = output[11:13]
        c = output[28:31]
        percent = int(a) / int(b)
        level = int(c)
        print(percent)
        print(level)
    except subprocess.SubprocessError:
        error = result.stderr
        print(f"Command execution failed with error: {error}")

    return percent, level


def get_signal_level_wireless():
    try:
        command1 = "cat /proc/net/wireless | grep wlan0 | awk '{print $3}'"  # 信号强度
        command2 = "cat /proc/net/wireless | grep wlan0 | awk '{print $4}'"  # 信号质量 单位dBm
        result1 = subprocess.run(command1, shell=True, stdout=PIPE, stderr=PIPE)
        result2 = subprocess.run(command2, shell=True, stdout=PIPE, stderr=PIPE)
        output1 = result1.stdout.strip()
        output2 = result2.stdout.strip()
        a = output1[0:2]
        c = output2[0:3]
        if len(a) == 0:
            a = 0
        if len(c) == 0:
            c = 0
        percent = int(a) / 70
        level = int(c)
        # print(percent)
        # print(level)
    except subprocess.SubprocessError:
        error1 = result1.stderr
        error2 = result2.stderr
        print(f"Command execution failed with error1: {error1}")
        print(f"Command execution failed with error2: {error2}")

    return percent, level


def get_picsize(data_path):
    # 列出指定路径下的所有文件和文件夹
    content1 = os.listdir(data_path)
    temp_path2 = os.path.join(data_path, content1[0])
    content2 = os.listdir(temp_path2)
    full_path = os.path.join(temp_path2, content2[0])
    picsize = os.path.getsize(full_path) / 1024 / 1000

    return picsize


def getenergy():
    with open('/sys/bus/i2c/drivers/ina3221x/6-0040/iio:device0/in_power0_input', 'r') as f:
        return float(f.read())


def inference(submodel1, img, util, size=224):
    time1 = time.time()
    transform = transforms.Compose([
        transforms.Resize(256),
        transforms.CenterCrop(size),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],
                             std=[0.229, 0.224, 0.225])
    ])
    img_tensor = transform(img).unsqueeze(0)

    with torch.no_grad():
        # output = model(img_tensor)
        out = submodel1(img_tensor)
        # print(out.shape)

        pre = out.detach().cpu().numpy()
        serialized_data = pickle.dumps(pre)
        size = sys.getsizeof(serialized_data)
        print('Serialized data size:', size, 'bytes')
        time2 = time.time()
        # 计算要上传的数据的大小（以字节为单位）
        # data_size = len(str(serialized_data).encode('utf-8'))
        # print('Serialized data size:', data_size, 'bytes')
        try:
            response = requests.post(f'http://{server_ip}:{server_port}/startEdge', data=serialized_data, timeout=5)
            output = response.text
            output = base64.b64decode(output)
            output = np.frombuffer(output, dtype='int64')
        except requests.exceptions.Timeout:
            output = [0, 0]
        except requests.exceptions.RequestException as e:
            output = [0, 0]
        except ValueError as e:
            output = [0, 0]

    time3 = time.time()

    # print("local time: {}".format(time2 - time1))
    # print("other time: {}".format(time3 - time2))
    util.settimelist(time2 - time1, time3 - time2)

    total_time = time3 - time1
    return output[0], total_time


def inference_1(model, img, util0, size=224):
    time1 = time.time()
    transform = transforms.Compose([
        transforms.Resize(256),
        transforms.CenterCrop(size),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],
                             std=[0.229, 0.224, 0.225])
    ])
    img_tensor = transform(img).unsqueeze(0)
    # img_tensor_ = img_tensor.squeeze(0)
    # img_array = (img_tensor_.detach().cpu().numpy() * 255).astype('uint8').transpose((1, 2, 0))
    # img_pil = Image.fromarray(img_array)
    # # img_pil = img_pil.rotate(-90)
    # img_pil.save('./data/test/n01558993/temp.py.jpg')

    with torch.no_grad():
        output = model(img_tensor)

    _, preds = torch.max(output, 1)

    preds = preds.detach().numpy()
    # print(preds)
    time2 = time.time()
    print("local reference time: {}".format(time2 - time1))
    util0.settimelist(time2 - time1)

    return preds[0], time2 - time1


def op_toExcel0(data, fileName):  # openpyxl库储存数据到excel
    wb = op.Workbook()  # 创建工作簿对象
    ws = wb['Sheet']  # 创建子表
    ws.append(['序号', '本地运行时间'])  # 添加表头
    for i in range(len(data)):
        d = data[i]["id"], data[i]["time1"]
        ws.append(d)  # 每次写入一行
    wb.save(fileName)
    return "Local writes sucess!"


def op_toExcel(data, fileName):  # openpyxl库储存数据到excel
    wb = op.Workbook()  # 创建工作簿对象
    ws = wb['Sheet']  # 创建子表
    ws.append(['序号', '本地运行时间', '剩余处理时间'])  # 添加表头
    for i in range(len(data)):
        d = data[i]["id"], data[i]["time1"], data[i]["time2"]
        ws.append(d)  # 每次写入一行
    # file_path = os.path.join(os.getenv('EXTERNAL_STORAGE'), fileName)

    wb.save(fileName)
    return "Local writes sucess!"


def op_toExcel2(data, fileName):  # openpyxl库储存数据到excel
    wb = op.Workbook()  # 创建工作簿对象
    ws = wb['Sheet']  # 创建子表
    ws.append(['序号', '总运行时间'])  # 添加表头
    for i in range(len(data)):
        d = data[i]["id"], data[i]["time1"]
        ws.append(d)  # 每次写入一行

    wb.save(fileName)
    return "Local writes sucess!"


def get_modeltype(model, position):
    path = ""
    info = [model, position]
    if model == "VGG16":
        path = "./model/vgg16/vgg16_pretrained_imagenet.pth"
        return VGG16.get_split_presubvgg16_mobile(path, info), 224
    elif model == "MobileNetV2":
        path = "./model/mobilenetv2/mobilenetv2_pretrained_imagenet.pth"
        return MobileNetV2.get_split_presubmobilebnetv2_mobile(path, info), 224
    elif model == "ResNet50":
        path = "./model/resnet50/resnet50_pretrained_imagenet.pth"
        return ResNet50.get_split_presubresnet50_mobile(path, info), 224
    # elif model == "AlexNet":
    #     path = "./model/alexnet/alexnet_pretrained.pth"
    #     return AlexNet.get_split_presubalexnet_mobile(path, info), 224
    # elif model == "ViT":
    #     path = "model/vit/vit_pretrained.pth"
    #
    # elif model == "MobileViT":
    #     path = "./model/mobilevit/mobilenvit_pretrained.pth"
    #     return MobileViT.get_split_presubmobilevit_mobile(path, info), 256
    else:
        print("No model type!")
        sys.exit(0)


def edge_computing(model, label_path, data_path):
    util2 = Util2()
    total_time = 0

    if model == "VGG16":
        util2, result, total_time = VGG16.edge_computing_vgg16(label_path, data_path, util2, server_ip, server_port)
    elif model == "MobileNetV2":
        util2, result, total_time = MobileNetV2.edge_computing_mobilenetv2(label_path, data_path, util2, server_ip,
                                                                           server_port)
    elif model == "ResNet50":
        util2, result, total_time = ResNet50.edge_computing_resnet50(label_path, data_path, util2, server_ip,
                                                                     server_port)
    # elif model == "AlexNet":
    #     util2, result, total_time = AlexNet.edge_computing_alexnet(label_path, data_path, util2, server_ip, server_port)
    # elif model == "SwinTransformer":
    #     path = ""
    # elif model == "MobileViT":
    #     util2, result, total_time = MobileViT.edge_computing_mobilevit(label_path, data_path, util2, server_ip, server_port)
    else:
        print("No model type!")
        sys.exit(0)

    # fileName = 'local_time_data.xlsx'
    # op_toExcel2(util2.gettimelist(), fileName)
    # print("本地结果已保存")

    return result, total_time


def local_computing(model, label_path, data_path, size):
    label_dict = json.load(open(label_path, "r"))

    inverted_dict = {value[0]: key for key, value in label_dict.items()}

    util0 = Util0()
    i = 1
    total = 0
    correct = 0
    for filefolder in os.listdir(data_path):
        answer = inverted_dict[filefolder]
        temp = os.path.join(data_path, filefolder)
        for filename in os.listdir(temp):
            img_path = os.path.join(temp, filename)
            print(img_path)
            img = Image.open(img_path)
            total += 1
            index, total_time = inference_1(model, img, util0, size)
            result = label_dict[str(index)][1]
            print("{} {} {}".format(i, result, filename))
            i += 1
            if int(answer) == index:
                correct += 1

    accuracy = correct / total
    # print("accuracy: {}".format(accuracy))

    # fileName = 'local_time_data.xlsx'
    # op_toExcel0(util0.gettimelist(), fileName)
    # print("本地结果已保存")

    return total_time


def split_computing(submodel1, label_path, data_path, size=224):
    label_dict = json.load(open(label_path, "r"))
    inverted_dict = {value[0]: key for key, value in label_dict.items()}
    util = Util()
    i = 1
    total = 0
    correct = 0
    for filefolder in os.listdir(data_path):
        answer = inverted_dict[filefolder]
        temp = os.path.join(data_path, filefolder)
        for filename in os.listdir(temp):
            img_path = os.path.join(temp, filename)
            print(img_path)
            img = Image.open(img_path)
            total += 1
            index, total_time = inference(submodel1, img, util, size)
            result = label_dict[str(index)][1]
            print("{} {} {}".format(i, result, filename))
            i += 1
            if int(answer) == index:
                correct += 1

    accuracy = correct / total
    print("accuracy: {}".format(accuracy))

    try:
        response = requests.post(f'http://{server_ip}:{server_port}/writeEXCEL', data=[], timeout=3)
        print(response.text)
    except requests.exceptions.Timeout:
        print("writeEXCEL timeout")
    except requests.exceptions.RequestException as e:
        print("writeEXCEL timeout")

    # fileName = 'local_time_data.xlsx'
    # op_toExcel(util.gettimelist(), fileName)
    # print("本地结果已保存")

    return accuracy, total_time


def main(modeltype, label_path, data_totalpath, buffer, epoches):
    for index in range(epoches):
        for childpath in os.listdir(data_totalpath):
            # 创建子进程对象并启动
            tegra_stats_process = TegraStatsProcess()
            tegra_stats_process.start()

            data_path = os.path.join(data_totalpath, childpath)
            transmission_rate = get_upload_speed()
            percent, strength = get_signal_level_wireless()  # 连接信号质量（百分比），信道强度（dbm）
            # picsize = get_picsize(data_path)
            if strength < -67:
                transmission_rate = random.randint(1, 5)
            state = [int(transmission_rate), round(percent, 2), strength]
            action_probs, action = agent.select_action(state)

            dict = {'VGG16': 17, 'AlexNet': 8, 'MobileNetV2': 21, 'ResNet50': 18, 'MobileViT': 13}
            if transmission_rate == 0:
                action = dict[modeltype]
            if transmission_rate < 1.2:
                action = random.randint(dict[modeltype] - 3, dict[modeltype])
            position = action
            info = [modeltype, position]

            try:
                response = requests.post(f'http://{server_ip}:{server_port}/getposition', json=info, timeout=2)
            except requests.exceptions.Timeout:
                position = dict[modeltype]
            except requests.exceptions.RequestException as e:
                position = dict[modeltype]

            if response.text != "success":
                print("modeltype and position send failed!")
                position = dict[modeltype]

            if position == 0:
                result, total_time = edge_computing(modeltype, label_path, data_path)
            elif position >= dict[modeltype]:
                localmodel, size = get_modeltype(modeltype, position)
                total_time = local_computing(localmodel, label_path, data_path, size)
            else:
                submodel1, size = get_modeltype(modeltype, position)
                accuracy, total_time = split_computing(submodel1, label_path, data_path, size)

            avg_energy = 0
            # 打印子进程的输出
            while not tegra_stats_process.stdout_queue.empty():
                output_line = tegra_stats_process.stdout_queue.get()
                # 使用正则表达式匹配模式并提取数值
                pattern = r'POM_5V_IN (\d+)/(\d+)'
                matches = re.findall(pattern, output_line)

                # 打印提取的数值
                for match in matches:
                    value1, value2 = match
                    # print(f"POM_5V_IN: {value1}/{value2}")
                # print(output_line)
                a, avg_energy = matches[-1]

            energy = int(avg_energy) * total_time / 60
            # 执行动作，并观察下一个状态、奖励和是否完成
            next_state, action, reward, average = agent.env_step_reward(state, action, total_time, energy,
                                                                        [modeltype, transmission_rate])
            print(state, "  ", action, "  ", reward)
            print("End-to-end Latency", total_time, "  ", "Energy Consumption", energy)

            buffer.add(state, action, reward)

            agent.upload.append(state[0])
            agent.link.append(state[1])
            agent.strength.append(state[2])
            # agent.size.append(state[3])
            agent.latency123.append(total_time)
            agent.energy.append(energy)

            episode_rewards = []
            log_probs = []
            total_reward = 0
            for i in range(1):
                log_prob = torch.log(action_probs[action])
                log_probs.append(log_prob)
                episode_rewards.append(reward)
                total_reward = total_reward + reward

                agent.x.append(agent.index)
                agent.y.append(total_reward)
                agent.point.append(action)
                agent.index = agent.index + 1
                total_reward = 0

            agent.episode_rewards.append(sum(episode_rewards))

            agent.update_policy(episode_rewards, log_probs, average)
            agent.replay(agent.policy_network, buffer, agent.optimizer, batch_size=20, avg=average)
            # 重启子进程
            tegra_stats_process.stop()
            tegra_stats_process.join()  # 等待子进程结束
            print("")
        print(agent.x)
        print(agent.y)
        agent.draw(agent.policy_network, modeltype)


if __name__ == '__main__':
    if torch.cuda.is_available():
        device = torch.device("cuda:{}".format(0))
    else:
        device = torch.device("cpu:0")

    modeltype = "VGG16"
    label_path = "./data/imagenet_class_index.json"
    data_path = "./data/test/"
    buffer = ReplayBuffer(buffer_size=50)
    epoches = 55
    main(modeltype, label_path, data_path, buffer, epoches)
